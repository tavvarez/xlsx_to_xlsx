import openpyxl
import time

def converter_planilhas_com_colunas_selecionadas(dicionario_planilhas_colunas):
    for arquivo, planilhas_colunas in dicionario_planilhas_colunas.items():
        # Carrega o arquivo XLSX original
        wb = openpyxl.load_workbook(arquivo)
        
        for nome_planilha, colunas in planilhas_colunas.items():
            # Seleciona a planilha atual
            planilha = wb[nome_planilha]
            
            # Cria um novo arquivo XLSX
            novo_wb = openpyxl.Workbook()
            nova_planilha = novo_wb.active
            
            # Copia as colunas desejadas para o novo arquivo XLSX
            for col_idx, coluna in enumerate(colunas, start=1):
                valores_coluna = planilha[coluna]
                for row_idx, celula in enumerate(valores_coluna, start=1):
                    nova_planilha.cell(row=row_idx, column=col_idx).value = celula.value
            
            # Salva o novo arquivo XLSX com um nome baseado no nome da planilha
            nome_saida = f'{arquivo}_{nome_planilha}_novo.xlsx'
            novo_wb.save(nome_saida)

# Exemplo de uso
dicionario_planilhas_colunas = {
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/SUZANO_MUCURI_BA.xlsx': {
        'MUCURI - BENS DE CONSUMO': ['C', 'F', 'I', 'H', 'K']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/SUZANO_MUCURI_BA_2.xlsx': {
        'PAPEL MUCURI': ['B', 'D', 'E', 'H']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/MARTINS.xlsx': {
        'MARTINS': ['C', 'D', 'E', 'F', 'G', 'I']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/JFM.xlsx': {
        'Base de Dados': ['A', 'B', 'C', 'D', 'G', 'I']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/SUZANO_BELEM.xlsx': {
        'BELEM - BENS DE CONSUMO': ['C', 'I', 'H', 'F', 'K']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/SUZANO_MARACANAU.xlsx': {
        'MARACANAU - BENS DE CONSUMO': ['C', 'I', 'H', 'F', 'J']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/SUZANO_VIANA_ES.xlsx': {
        'VIANA - BENS DE CONSUMO': ['C', 'I', 'H', 'F', 'K']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/SUZANO_IMPERATRIZ.xlsx': {
        'IMPERATRIZ - BENS DE CONSUMO': ['C', 'I', 'H', 'F', 'K']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/SUZANO_LIMEIRA.xlsx': {
        'PAPEL LIMEIRA': ['B', 'E', 'D', 'H']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/NATVILLE.xlsx': {
        'NATVILLE': ['C', 'D', 'G', 'F', 'E', 'I']
    },
    'C:/Users/Bialog-006/desktop/PlanilhasBOT/INTECOM_ALHANDRA.xlsx': {
        'INTECOM': ['B', 'C', 'D', 'E', 'F']
    },
}

start_time = time.time()

converter_planilhas_com_colunas_selecionadas(dicionario_planilhas_colunas)

end_time = time.time()
execution_time = end_time - start_time
print(f"Tempo de execução: {round(execution_time, 3)} segundos")
